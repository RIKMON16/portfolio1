weight_lbs = input('What is your weight? (lbs): ')
weight_kgs = int(weight_lbs) // 2.25
print(weight_kgs)
# single quote example
course = "Python's Course for Beginners"
print(course)
# double quote example "Beginners"
course = 'Python Course for "Beginners"'
print(course)
# triple quote in case the string is large in length. We need ''' to start and ''' to end string
course = '''
Hi John,
We have received your email. 
We shall get back quickly.
Thank You Support Team.
'''
print(course)
# to check the value of 0 for value p and for reverse use negative value start from -1 to get s
# to get more values use [0:4] it will show the chars from 0 (P) 1(y) 2(t) 3(h) but not the 4th char, if you do not
# provide the end char in list python will reflect all char [0:] to get all chars
course = 'Python for Beginners'
object = course[:]
print(object)

name = 'Rishipal'
print(name[1:-1])

# formatted strings
# ref name msg is formatted string
first = 'Rishi'
last = 'Kalsi'
message = first + '[' + last + '] is a coder'
msg = f'{first} {[last]} is a coder'
print(msg)

# concatenation In formal language theory and computer programming,
# string concatenation is the operation of joining character strings end-to-end.
# For example, the concatenation of "snow" and "ball" is "snowball"

# string methods
# len stands for counting the length of words. with len function we can fix the number of chars of input from user
# it is a general purpose function can also be used to count the number of items in a list
course = 'Python for Beginners'
object = course[:]
print(len(object))

# use . operator to get access to different string functions example as follows
course = 'Python for Beginners'
object = course[:]
print(object.upper())

course = 'Python for Beginners'
object = course[:]
print(object.find('P'))

course = 'Python for Beginners'
object = course[:]
print(object.replace('Beginners', 'Absolute Beginners'))

# to find something in a code use in operator as in below example
# this is a expression which will represent Bool value
# if Python is mentioned with capital letter and u mention it with lower , Bool will be false
course = 'Python for Beginners'
object = course[:]
print('Python' in object)

# Airthmetic Operations

x = 10
y = 3
z = x + y
print(z)

# augmented assignment operator
x = 10
x += 3
print(x)

# operator precedence
# why ans is 16 beacuse multiply is condiered higher in value than add, so the
# command will follow first by multiply and then add

x = 10 + 3 * 2 ** 2
print(x)

# order of operations as per above example
# exponentiation to raise one qty with the power of another 2 ** 3. here 2 is qty and 3 is power
# multiplication or division, addittion or subtraction
# to prioritise use parenthesis ()

x = (2 + 3) * 10 - 3
print(x)

# Math Functions. if you have to perform math then you need to import math module
# to learn more refer google python3 math module

import math

print(math.floor(2.9))
print(math.ceil(2.9))

# if Statements. They allow us to make a program based on these conditions computer is able to think and ans
# so if some conditions are true we are going to do ceratin thing or if false we are going to do another

is_hot = False
is_cold = False

if is_hot:
    print("It's a hot day")
    print("Drink Plenty of Water")
elif is_cold:
    print("It's a cold day")
    print("Wear warm clothes")
else:
    print("It's a lovely day")
print("Enjoy your day")

# excercise: Imagine cost of house is $1M. if the buyer has good credit he need to pay 10%,
# otherwise they need to put down 20%. print down payment

is_goodcredit = False
is_avgcredit = False
if is_goodcredit:
    print("Cost of house is $1M")
    print("as you have good credit history. you need to pay 10% of $1M in advance")
elif is_avgcredit:
    print("Cost of house is $1M")
    print("as you have avg credit history. you need to pay 20% of $1M in advance")
else:
    print("Cost of house is $1M")
    print("as you have bad credit history. we are unable to process loan")

# solution as per mosh.. i should have thought it in a better way

price = 1000000
is_goodcredit = False
is_avgcredit = False

if is_goodcredit:
    down_payment = 0.1 * price
    print(f"as you have good credit history. you need to pay 10% down payment: ${down_payment} in advance")
elif is_avgcredit:
    down_payment = 0.2 * price
    print(f"as you have avg credit history. you need to pay 20% down payment: ${down_payment} in advance")
else:
    down_payment = 0.5 * price
    print(f"as you have bad credit history. you need to pay 50% down payment: ${down_payment} in advance")

# Logical Operators - AND is a logical operator is used to combine two statements
# AND operator both conditions should be true
# OR operator one condition should be true

has_high_income = True
has_good_credit = True

if has_high_income and has_good_credit:
    print("Eligible for loan")

has_high_income = False
has_good_credit = True

if has_high_income or has_good_credit:
    print("Eligible for loan")

has_criminal_record = True
has_good_credit = True

if has_good_credit and not has_criminal_record:
    print("Eligible for loan")
else:
    print("Not eligible for loan")

# comparison operators

temperature = 30

if temperature > 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

# project: Weight converter

weight = int(input('What is your weight? : '))
unit = input('(L)bs or (K)g: ')
if unit.upper() == "L":
    converted = weight * 0.45
    print(f" You are {converted} kilos")
else:
    converted = weight // 0.45
    print(f"You are {converted} pounds")

# While Loops important - y loop is the ingredient to built games
i = 1
while i <= 5:
    print(i)
    i = i + 1
print("Done")

# in this loop till the time i reaches 5 loop will continue as our
# condition is i is less than or equal to 5 once i surpasses 5 it will end loop and print done

# now i got the pyramid of mario
i = 1
while i <= 5:
    print('#' * i)
    i = i + 1
print("Done")

# Guessing Game
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    guess = int(input('Guess: '))
    guess_count += 1
    if guess == secret_number:
        print('You Won!')
        break
    else:
        print('Sorry you failed')

# Car Game

# when press help screen will show start - to start the car. stop - to stop the car. quit - to exit.
# if i type anything else it should say I don't understand that... on typing start - must say Car started.... Ready to go
# when type quit our program is terminated
# to remove duplication of lower command write .lower after input parentesis

command = ""
while True:
    command = input("> ").lower()
    if command == "start":
        print("Car Started.... Ready to go ken")
    elif command == "stop":
        print("Car Stoped..")
    elif command == "help":
        print("""
start - to start the car
stop - to stop the car
quit - to end
        """)
    elif command == "quit":
        break
    else:
        print("Sorry. I don't understand")
# if user type quit right away the line sorry. i don't understand will display to sort this problem we need another
# elif command with quit and after that enter break

command = ""
started = False
while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car Already Start...")
        else:
            started = True
            print("Car Started.... Ready to go ken")
    elif command == "stop":
        if not started:
            print("Car is already stopped")
        else:
            started = False
            print("Car Stoped..")
    elif command == "help":
        print("""
start - to start the car
stop - to stop the car
quit - to end
        """)
    elif command == "quit":
        break
    else:
        print("Sorry. I don't understand")

# For Loops
for item in 'Python':
    print(item)

for item in ['John', 'Mosh', 'Rishi']:
    print(item)

# python have builtin function of range which is helpful to create list of 1 to n number
# also you can make the range from a starting point of your choice plus you can add a gap

for item in range(20):
    print(item)

for item in range(3, 20, 2):
    print(item)

# excercise sum the cost of shopping cart using for loop prices = [10,20,30]

prices = [10, 20, 30]

total = 0
for price in prices:
    total += price
print(f"Total: {total}")

# Nested Loops means adding loop inside of another loop
# with this we can easily generate cordinates, which are acquired as you know is  a combination of x and y

for x in range(5):
    for y in range(4):
        print(f'({x},{y})')

# how to make it look like f
# simple way to do it until for loop its good then type print('x' * x_count) it will display F

numbers = [5, 2, 5, 2, 2]
for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'x'
    print(output)

# to print L
numbers = [2, 2, 2, 2, 7, 7]
for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'x'
    print(output)

# lists
names = ['John', 'marry', 'bob', 'marley']
print(names)

# write a program to find the largest number in the list

numbers = [1, 2, 3, 4, 5, 6, 4]
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)

# 2D lists - 2 dimensional lists are very powerful in data science and machine learning

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
print(matrix[2][2])
# 0 stands for list 1 and 1 stands for item number in the list. similary you can select list 2 which 3rd list and item 2

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
for row in matrix:
    for item in row:
        print(item)

# list methods
# append will add number at the end of the list
# to insert number at different location use method insert
# remove method to remove any number / item from the list
# to remove everything from list use clear
# pop will remove last digit of the list
numbers = [1, 4, 7, 9, 13]
numbers.append(30)
print(numbers)

numbers = [1, 4, 7, 9, 13]
numbers.insert(2, 8)
print(numbers)

# write a program to remove duplicates in the list

numbers = [1, 3, 3, 4, 4, 7, 9, 10, 13]
uniques = []
for number in numbers:
    if number not in uniques:
        uniques.append(number)
print(uniques)

# Tuples are similar as lists to be used to save list of items but unlike lists they can;t be modified or add new item,
# remove existing item. list is definded in [brackets] and tuple is defined in (parenthses)

numbers = (1, 4, 7, 9, 13)
print(numbers)

# unpacking

coordinates = (1, 2, 3)
x, y, z = coordinates
print(x)
print(y)
print(z)

# Dictionaries - customer data. to define dic you need to mention data in curly braces {}
# . get method to not to be yelled by computer

customer = {
    "name": "john smith",
    "age": "40",
    "is_verified": True
}
print(customer.get("name"))

phone = input("Phone: ")
digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"
}
output = ""
for ch in phone:
    output += digits_mapping.get(ch, "!") + " "
print(output)


# emoji converter. to get emoji character in mac press control and space
# dont have emoji keyboard

# message = input(">")
#   words = message.split(" "):
#  emojis = {
#     ":)": "😀",
#    ":(": "😞"
# }
# output = ""
# for word in words:
#   output += emojis.get(word, word) + " "
# print(output)

# Functions
def greet_user(first_name, last_name):
    print(f'HI {first_name} {last_name}!')
    print('Welcome on board')


print("Start")
greet_user("John", "Smith")
greet_user("Rishi", "Kalsi")
print("Finish")


# Parameters are giving instructions to functions likewise given to greet_user as if first and last_name

# keyword arguments. example as follows

def greet_user(first_name, last_name):
    print(f'HI {first_name} {last_name}!')
    print('Welcome on board')


print("Start")
greet_user(last_name="John", first_name="Smith")
print("Finish")


# return statement

def square(number):
    return number * number


print(square(5))


# Creating a reusable function

def emoji_converter(message):
    words = message.split(" ")
    emojis = {
        ":)": "😀",
        ":(": "😞"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">")
print(emoji_converter(message))

# Exceptions
try:
    age = int(input('Age: '))
    income = 20000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print('Age cannot be 0.')
except ValueError:
    print('Invalid Value')


# Comments - we can use this to communicate to make reminders

# Classes in python. Very important for programming.
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()

point2 = Point()
point2.x = 1
print(point2.x)


# Constructors
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point(10, 20)
print(point.x)


# Exercise Person make talk and walk

class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi, I am {self.name}")


john = Person("John Smith")
john.talk()

jazz = Person("Jasmeet Kaur")
jazz.talk()


# Inheritance
class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print("bark")


class Cat(Mammal):
    def be_annoying(self):
        print("be annoying")


dog1 = Dog()
dog1.bark()

cat1 = Cat()
cat1.be_annoying()

# Modules
import converter

print(converter.kg_to_lbs(100))

# Exercise
from utils import find_max

numbers = [10, 3, 4, 8, 20, 13, 21]
maximum = find_max(numbers)
print(maximum)

# Packages
# from ecommerce.shipping import calc_shipping

from ecommerce import shipping

shipping.calc_shipping()

# Generating Random Values
# in this chapater tutor is telling to use inbuilt modules of python. to find them search on google python3 module index

import random

for i in range(3):
    print(random.random())

# use randint to find random numbers between 2 numbers
import random

for i in range(5):
    print(random.randint(5, 20))

# another example
import random

members = ['Rishi', 'Jas', 'Nisha', 'Kunal', 'Vick', 'love', 'rupi']
leader = random.choice(members)
print(leader)

# create program dice in dice create tuple in roll
import random

dice1 = ('1', '2', '3', '4', '5', '6')
dice2 = ('1', '2', '3', '4', '5', '6')
roll = random.choice(dice1)
print(roll)

import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())

# Working with directories
# Absolute Path - is the path from the root of the hard disk. like drives on the hard drive
# Relative path - is like ecommerce directory which is a part of this file


from pathlib import Path

path = Path("ecommerce")
print(path.exists())

# it will return Bool value
# to make directory
from pathlib import Path

path = Path("email")
print(path.mkdir())

# to remove
from pathlib import Path

path = Path("email")
print(path.rmdir())

# Automation with Python. Working on Excel Spreadsheets

import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
                   min_row=2, max_row=sheet.max_row,
                   min_col=4, max_col=4)

chart = BarChart3D()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions.xlsx')

# better way to write code as a professional whereever transactions is written we can change the name
# and use the same function to work on number of files
import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference


def process_workbook(filename):
    wb = xl.load_workbook(transactions.xlsx)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2, max_row=sheet.max_row,
                       min_col=4, max_col=4)

    chart = BarChart3D()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(transactions1.xlsx)

# Machine learing or Artifical Intelligence
# Steps - 1. Import the Data (CSV. ETC), 2. Clean the Data (delete duplicates),
# 3. Split the data into Training / Test sets, 4. Create a Model, 5. Train the Model,
# 6. Make Predications, 7. Evalute & Improve

# Libraries & Tools
# 1. Numpy - which provides multidimenssional array. 2. Pandas - data analysis library, provide concept called data frame.
# 3. MatPlotLib - which is a 2 diminsional library for creating graphs on plot
# 4. Scikit-Learn - most popular all common algorithums like decision trees, neural networks.
# for machine learning jupyter is the enviornment which is used

# Importing Data Set to Jupyter from CSV format. Download data set from kaggle.com